# ──────────────────────────────────────────
# IMPORTS
# ──────────────────────────────────────────

import io
from datetime import date
import secrets
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Emprestimo, Livro, Usuario
import json
import random
from collections import defaultdict
from datetime import date
from .reservas import (
    criar_reserva, aceitar_reserva, recusar_reserva,
    confirmar_disponibilidade, confirmar_retirada,
    cancelar_reserva_usuario, verificar_devolucao_com_reserva,
    ReservaError,
)   
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.db.models import Exists, OuterRef, Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.views.decorators.http import require_POST
from .gmail import enviar_email
from rest_framework import filters, viewsets
from .models import Emprestimo, Exemplar, Lista, Livro, LivroLido, Reserva, Usuario
from .serializers import LivroSerializer
from django.core.paginator import Paginator
from .models import Emprestimo, Exemplar, Lista, Livro, LivroLido, Reserva, Usuario, Turma, Cargo, ABAS_DISPONIVEIS

from calendar import monthrange
from django.db.models import Count
from django.db.models.functions import TruncWeek, TruncMonth
from datetime import date, datetime, timedelta
from .models import Notificacao

# ──────────────────────────────────────────
# AUTENTICAÇÃO
# ──────────────────────────────────────────

@ensure_csrf_cookie
def pagina_login(request):
    if request.user.is_authenticated:
        return redirect('inicio')
    return render(request, 'registration/login.html')


def verificar_matricula(request):
    if request.method != 'POST':
        return JsonResponse({'erro': 'Método inválido'}, status=405)

    data = json.loads(request.body)
    matricula = data.get('matricula', '').strip()

    try:
        usuario = Usuario.objects.get(matricula=matricula)
        return JsonResponse({
            'existe':         True,
            'primeiro_acesso': usuario.primeiro_acesso,
            'nome':           usuario.get_full_name() or usuario.username,
        })
    except Usuario.DoesNotExist:
        return JsonResponse({'existe': False})


def fazer_login(request):
    if request.method != 'POST':
        return JsonResponse({'erro': 'Método inválido'}, status=405)

    data = json.loads(request.body)
    matricula = data.get('matricula', '').strip()
    senha = data.get('senha', '').strip()
    lembrar = bool(data.get('lembrar', False))

    try:
        usuario = Usuario.objects.get(matricula=matricula)
    except Usuario.DoesNotExist:
        return JsonResponse({'erro': 'Matrícula não encontrada.'}, status=404)

    user = authenticate(request, username=usuario.username, password=senha)
    if user is not None:
        login(request, user)

        if lembrar:
            # Sessão persiste por 30 dias, mesmo fechando o navegador.
            request.session.set_expiry(60 * 60 * 24 * 30)
        else:
            # Expira assim que o navegador for fechado.
            request.session.set_expiry(0)

        destino = '/biblioteca/emprestimos/' if user.is_bibliotecario else '/biblioteca/'
        return JsonResponse({'sucesso': True, 'redirect': destino})

    return JsonResponse({'erro': 'Senha incorreta.'}, status=401)


def primeiro_acesso(request):
    if request.method != 'POST':
        return JsonResponse({'erro': 'Método inválido'}, status=405)

    data = json.loads(request.body)
    matricula = data.get('matricula', '').strip()
    senha = data.get('senha', '').strip()
    email = data.get('email', '').strip()

    try:
        usuario = Usuario.objects.get(matricula=matricula)
    except Usuario.DoesNotExist:
        return JsonResponse({'erro': 'Matrícula não encontrada.'}, status=404)

    if not usuario.primeiro_acesso:
        return JsonResponse({'erro': 'Essa conta já foi ativada.'}, status=400)
    if len(senha) < 8:
        return JsonResponse({'erro': 'Senha deve ter pelo menos 8 caracteres.'}, status=400)
    if not email:
        return JsonResponse({'erro': 'Informe um e-mail.'}, status=400)

    usuario.set_password(senha)
    usuario.email = email
    usuario.primeiro_acesso = False
    usuario.save()

    user = authenticate(request, username=usuario.username, password=senha)
    login(request, user)
    return JsonResponse({'sucesso': True, 'redirect': '/biblioteca/'})


def fazer_logout(request):
    logout(request)
    return redirect('login')


def cadastro(request):
    matricula = request.GET.get('matricula', '').strip()
    if not matricula:
        return redirect('login')

    try:
        usuario = Usuario.objects.get(matricula=matricula)
    except Usuario.DoesNotExist:
        return redirect('login')

    if not usuario.primeiro_acesso:
        return redirect('login')

    iniciais = ''
    partes = usuario.get_full_name().split()
    if len(partes) >= 2:
        iniciais = partes[0][0].upper() + partes[-1][0].upper()
    elif partes:
        iniciais = partes[0][0].upper()

    return render(request, 'registration/cadastro.html', {
        'matricula': matricula,
        'nome':      usuario.get_full_name(),
        'serie':     usuario.serie or '',
        'iniciais':  iniciais,
    })


def recuperar_senha(request):
    return render(request, 'registration/recuperar_senha.html')

def _mascarar_email(email):
    try:
        user, dominio = email.split('@')
    except ValueError:
        return '***'
    prefixo = user[0] + '*' if len(user) <= 2 else user[:2] + '*' * (len(user) - 2)
    return f'{prefixo}@{dominio}'


@require_POST
def recuperar_buscar_contato(request):
    data = json.loads(request.body)
    matricula = data.get('matricula', '').strip()

    try:
        usuario = Usuario.objects.get(matricula=matricula)
    except Usuario.DoesNotExist:
        return JsonResponse({'existe': False, 'erro': 'Matrícula não encontrada.'}, status=404)

    if usuario.primeiro_acesso:
        return JsonResponse({
            'existe': False,
            'erro': 'Essa matrícula ainda não tem senha cadastrada. Use "Criar minha senha" na tela de login.',
        }, status=400)

    if not usuario.email:
        return JsonResponse({
            'existe': False,
            'erro': 'Nenhum e-mail cadastrado para esta matrícula. Fale com a bibliotecária.',
        }, status=400)

    return JsonResponse({'existe': True, 'email_mascarado': _mascarar_email(usuario.email)})

@require_POST
def enviar_codigo(request):
    data = json.loads(request.body)
    matricula = data.get('matricula', '').strip()
    fluxo = data.get('fluxo', 'cadastro').strip()

    try:
        usuario = Usuario.objects.get(matricula=matricula)
    except Usuario.DoesNotExist:
        return JsonResponse({'erro': 'Usuário não encontrado.'}, status=404)

    if fluxo == 'recuperacao':
        # Recuperação: o e-mail vem do banco, nunca do cliente.
        email = usuario.email
        if not email:
            return JsonResponse({'erro': 'Nenhum e-mail cadastrado para esta matrícula.'}, status=400)
    else:
        # Cadastro (primeiro acesso): o e-mail ainda não existe no banco.
        email = data.get('email', '').strip()
        if not email:
            return JsonResponse({'erro': 'Informe um e-mail.'}, status=400)

    codigo = str(random.randint(100000, 999999))
    cache.set(f"codigo_verificacao_{matricula}", codigo, timeout=600)

    from .gmail import enviar_codigo_verificacao
    enviar_codigo_verificacao(email, codigo)

    return JsonResponse({'sucesso': True})


@require_POST
def verificar_codigo(request):
    data = json.loads(request.body)
    matricula = data.get('matricula', '').strip()
    codigo_digitado = data.get('codigo', '').strip()
    fluxo = data.get('fluxo', 'cadastro').strip()

    cache_key = f"codigo_verificacao_{matricula}"
    codigo_salvo = cache.get(cache_key)

    if not codigo_salvo:
        return JsonResponse({'erro': 'Código expirado. Solicite um novo.'}, status=400)
    if codigo_digitado != codigo_salvo:
        return JsonResponse({'erro': 'Código incorreto.'}, status=400)

    cache.delete(cache_key)

    if fluxo == 'recuperacao':
        # Gera um token de curta duração para autorizar a troca de senha na próxima tela.
        token = secrets.token_urlsafe(32)
        cache.set(f"reset_token_{matricula}", token, timeout=600)
        return JsonResponse({'sucesso': True, 'reset_token': token})

    return JsonResponse({'sucesso': True})

@require_POST
def redefinir_senha_recuperacao(request):
    data = json.loads(request.body)
    matricula = data.get('matricula', '').strip()
    token = data.get('token', '').strip()
    nova_senha = data.get('senha_nova', '').strip()

    if not matricula or not token or not nova_senha:
        return JsonResponse({'erro': 'Dados incompletos.'}, status=400)
    if len(nova_senha) < 8:
        return JsonResponse({'erro': 'A nova senha deve ter pelo menos 8 caracteres.'}, status=400)

    cache_key = f"reset_token_{matricula}"
    token_salvo = cache.get(cache_key)
    if not token_salvo or token_salvo != token:
        return JsonResponse({'erro': 'Sessão de redefinição expirada. Solicite um novo código.'}, status=400)

    try:
        usuario = Usuario.objects.get(matricula=matricula)
    except Usuario.DoesNotExist:
        return JsonResponse({'erro': 'Usuário não encontrado.'}, status=404)

    usuario.set_password(nova_senha)
    usuario.save()
    cache.delete(cache_key)

    return JsonResponse({'sucesso': True, 'redirect': '/biblioteca/login/'})

def verificar_codigo_page(request):
    matricula = request.GET.get('matricula', '').strip()
    contato = request.GET.get('contato', '').strip()
    fluxo = request.GET.get('fluxo', 'cadastro').strip()

    if not matricula or not contato:
        return redirect('login')

    return render(request, 'registration/verificar_codigo.html', {
        'matricula': matricula,
        'contato':   contato,
        'fluxo':     fluxo,
    })

def nova_senha_page(request):
    matricula = request.GET.get('matricula', '').strip()
    if not matricula:
        return redirect('login')
    return render(request, 'registration/nova_senha.html', {'matricula': matricula})

# ──────────────────────────────────────────
# PÁGINAS PRINCIPAIS
# ──────────────────────────────────────────

def inicio(request):
    hoje = date.today()
    usuario = request.user if request.user.is_authenticated else None

    CORES = ['#1e5aa8', '#7c3aed', '#059669', '#b45309', '#dc2626', '#0891b2', '#be185d']
    ICONES_LISTA = ['📚', '⭐', '🎯', '📖', '🔖', '💡', '🏆']
    ROTS = [-6, -2, 4]

    exemplares_emprestados_ids = Emprestimo.objects.filter(
        data_devolucao_real__isnull=True
    ).values_list('exemplar_id', flat=True)

    livros_sem_disponivel = Exemplar.objects.filter(
        id_exemplar__in=exemplares_emprestados_ids
    ).values_list('livro_id', flat=True)

    sugestoes = Livro.objects.exclude(
        id_livro__in=livros_sem_disponivel
    ).order_by('?')[:10]

    listas = []
    emprestimos_recentes = []
    prazo_urgente = None

    if usuario:
        listas_qs = Lista.objects.filter(usuario=usuario).prefetch_related('livros')
        for l in listas_qs:
            livros = l.livros.all()[:3]
            l.livros_preview = [
                {
                    'titulo':     livro.titulo,
                    'capa_url':   livro.capa_url or '',
                    'cover_from': '#1e5aa8',
                    'cover_to':   '#0b1526',
                    'rot':        ROTS[i % len(ROTS)],
                }
                for i, livro in enumerate(livros)
            ]
            l.icone  = ICONES_LISTA[l.pk % len(ICONES_LISTA)]
            l.cor    = CORES[l.pk % len(CORES)]
            l.vis    = 'privada'
            l.criada = l.criada_em
            listas.append(l)

        qs = Emprestimo.objects.filter(
            usuario=usuario,
            data_devolucao_real__isnull=True,
        ).select_related('exemplar__livro').order_by('data_devolucao_prevista')[:3]

        for emp in qs:
            delta = (emp.data_devolucao_prevista - hoje).days
            emp.livro             = emp.exemplar.livro
            emp.dias_restantes    = max(delta, 0)
            emp.dias_usados       = max((hoje - emp.data_emprestimo).days, 0)
            emp.dias_total        = 15
            emp.percentual_usado  = min(int((emp.dias_usados / 15) * 100), 100)
            emp.urgente           = 0 <= delta <= 3 or delta < 0
            emp.renovacoes_usadas = 0
            emp.renovacoes_max    = 1
            emp.cor_avatar        = CORES[emp.pk % len(CORES)]
            emprestimos_recentes.append(emp)

        urgentes = [e for e in emprestimos_recentes if e.urgente]
        if urgentes:
            prazo_urgente = urgentes[0]
            prazo_urgente.data_devolucao = prazo_urgente.data_devolucao_prevista

    return render(request, 'biblioteca/inicio.html', {
        'sugestoes':            sugestoes,
        'listas':               listas,
        'emprestimos_recentes': emprestimos_recentes,
        'total_lidos':          0,
        'total_emprestados':    len(emprestimos_recentes),
        'dia_semana':           hoje.strftime('%A'),
        'data_hoje':            hoje.strftime('%d/%m/%Y'),
        'prazo_urgente':        prazo_urgente,
    })


def detalhes_livro(request, livro_id):
    livro = get_object_or_404(Livro, id_livro=livro_id)
    disponivel = livro.esta_disponivel()

    ja_lido = (
        request.user.is_authenticated and
        LivroLido.objects.filter(usuario=request.user, livro=livro).exists()
    )

    return render(request, 'biblioteca/detalhes_livro.html', {
        'livro':     livro,
        'disponivel': disponivel,
        'ja_lido':   ja_lido,
    })


def explorar(request):
    from django.core.paginator import Paginator

    # ── Parâmetros da URL ──────────────────────────────────────
    page_number  = request.GET.get('page', 1)
    filtro_cat   = request.GET.get('cat', '')       # ex: ?cat=Romance
    filtro_prat  = request.GET.get('prat', '')      # ex: ?prat=A
    filtro_disp  = request.GET.get('disp', '')      # ex: ?disp=disponivel
    filtro_ordem = request.GET.get('ordem', 'az')   # ex: ?ordem=za

    # ── Subquery de disponibilidade ────────────────────────────
    tem_disponivel = Exemplar.objects.filter(
        livro=OuterRef('pk'),
        status='disponivel'
    )

    # ── QuerySet base ──────────────────────────────────────────
    livros = Livro.objects.annotate(disponivel=Exists(tem_disponivel))

    # ── Filtros opcionais ──────────────────────────────────────
    if filtro_cat:
        livros = livros.filter(categoria=filtro_cat)

    if filtro_prat:
        livros = livros.filter(prateleira=filtro_prat)

    if filtro_disp == 'disponivel':
        livros = livros.filter(disponivel=True)
    elif filtro_disp == 'indisponivel':
        livros = livros.filter(disponivel=False)

    # ── Ordenação ──────────────────────────────────────────────
    ordem_map = {
        'az':       'titulo',
        'za':       '-titulo',
        'recentes': '-data_cadastro',
    }
    livros = livros.order_by(ordem_map.get(filtro_ordem, 'titulo'))

    # ── Paginação ──────────────────────────────────────────────
    paginator = Paginator(livros, 24)
    page_obj  = paginator.get_page(page_number)

     # ── Janela de páginas para o template ─────────────────────
    page_num      = page_obj.number
    total_pages   = paginator.num_pages
    vizinhos      = 2

    inicio = max(page_num - vizinhos, 2)
    fim    = min(page_num + vizinhos, total_pages - 1)
    pag_janela = list(range(inicio, fim + 1))

    # ── Listas para os filtros (sempre completas, sem filtro) ──
    categorias = (
        Livro.objects
        .exclude(categoria__isnull=True).exclude(categoria='')
        .values_list('categoria', flat=True)
        .distinct().order_by('categoria')
    )
    prateleiras = (
        Livro.objects
        .exclude(prateleira__isnull=True).exclude(prateleira='')
        .values_list('prateleira', flat=True)
        .distinct().order_by('prateleira')
    )

    return render(request, 'biblioteca/explorar.html', {
        'page_obj':     page_obj,
        'categorias':   categorias,
        'prateleiras':  prateleiras,
        'total_livros': paginator.count,
        'filtro_cat':   filtro_cat,
        'filtro_prat':  filtro_prat,
        'filtro_disp':  filtro_disp,
        'filtro_ordem': filtro_ordem,
        'pag_janela':   pag_janela,
    })
@login_required
def prazos(request):
    hoje = date.today()

    CORES = ['#1e5aa8', '#7c3aed', '#059669', '#b45309', '#dc2626', '#0891b2', '#be185d']

    emprestimos_qs = Emprestimo.objects.filter(
        usuario=request.user,
        data_devolucao_real__isnull=True,
    ).select_related('exemplar__livro').order_by('data_devolucao_prevista') \
        if request.user.is_authenticated else Emprestimo.objects.none()

    total_ok = total_breve = total_atrasados = 0

    for emp in emprestimos_qs:
        delta = (emp.data_devolucao_prevista - hoje).days
        emp.vence_em_breve   = 0 <= delta <= 3
        emp.dias_restantes   = max(delta, 0)
        emp.dias_atraso      = abs(delta) if delta < 0 else 0
        emp.dias_usados      = max((hoje - emp.data_emprestimo).days, 0)
        emp.percentual_usado = min(int((emp.dias_usados / 15) * 100), 100)
        emp.cor_avatar       = CORES[emp.pk % len(CORES)]

        if emp.esta_atrasado():
            total_atrasados += 1
        elif emp.vence_em_breve:
            total_breve += 1
        else:
            total_ok += 1

    historico = Emprestimo.objects.filter(
        usuario=request.user,
        data_devolucao_real__isnull=False,
    ).order_by('-data_devolucao_real') if request.user.is_authenticated else Emprestimo.objects.none()

    historico_count  = historico.count()
    ultimo_devolvido = historico.first().data_devolucao_real if historico_count > 0 else None

    return render(request, 'biblioteca/prazo.html', {
        'emprestimos':      emprestimos_qs,
        'total_ativos':     emprestimos_qs.count(),
        'total_ok':         total_ok,
        'total_breve':      total_breve,
        'total_atrasados':  total_atrasados,
        'historico_count':  historico_count,
        'ultimo_devolvido': ultimo_devolvido,
    })


def configuracoes(request):
    return render(request, 'biblioteca/configuracoes.html')


# ──────────────────────────────────────────
# ACERVO / ABAS
# ──────────────────────────────────────────

def acervo(request):
    q = request.GET.get('q', '')
    if q:
        livros = Livro.objects.filter(titulo__icontains=q) | Livro.objects.filter(autor__icontains=q)
    else:
        livros = Livro.objects.all()[:50]
    return render(request, 'biblioteca/acervo_busca.html', {'livros': livros, 'q': q})

def _acervo_counts(user):
    if not user.is_authenticated:
        return {'listas_count': 0, 'lidos_count': 0, 'reservas_count': 0}
    return {
        'listas_count':  Lista.objects.filter(usuario=user).count(),
        'lidos_count':   LivroLido.objects.filter(usuario=user).count(),
        'reservas_count': Reserva.objects.filter(
            usuario=user, status__in=['pendente', 'fila', 'aceita', 'aguardando_retirada']
        ).count(),
    }

@login_required
def lista(request):
    listas_qs = Lista.objects.filter(usuario=request.user).prefetch_related('livros') \
        if request.user.is_authenticated else []

    listas = []
    for l in listas_qs:
        livros = l.livros.all()[:3]
        l.livros_preview = [
            {
                'titulo':     livro.titulo,
                'capa_url':   livro.capa_url or '',
                'cover_from': '#1e5aa8',
                'cover_to':   '#0b1526',
            }
            for livro in livros
        ]
        listas.append(l)

    return render(request, 'biblioteca/acervo.html', {'aba': 'lista', 'listas': listas, **_acervo_counts(request.user),})

@require_POST
@login_required
def reservar_livro(request, livro_id):
    livro = get_object_or_404(Livro, id_livro=livro_id)
    try:
        reserva, entrou_fila = criar_reserva(request.user, livro)
    except ReservaError as e:
        return JsonResponse({'erro': str(e)}, status=400)

    if entrou_fila:
        return JsonResponse({
            'sucesso': True,
            'fila': True,
            'posicao_fila': reserva.posicao_fila,
            'mensagem': (
                f'Este livro já foi reservado por outro aluno. Você entrou na fila '
                f'na posição {reserva.posicao_fila} de no máximo 3.'
            ),
        })
    return JsonResponse({
        'sucesso': True,
        'fila': False,
        'mensagem': 'Reserva enviada! Aguarde a aprovação da bibliotecária.',
    })


@require_POST
@staff_member_required
def reserva_confirmar_retirada(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    try:
        emprestimo = confirmar_retirada(reserva)
    except ReservaError as e:
        return JsonResponse({'erro': str(e)}, status=400)
    return JsonResponse({'sucesso': True, 'emprestimo_id': emprestimo.pk})

@login_required
def lidos(request):
    livros_lidos = []
    lidos_count = 0

    if request.user.is_authenticated:
        qs = LivroLido.objects.filter(
            usuario=request.user
        ).select_related('livro')

        lidos_count = qs.count()

        for registro in qs:
            registro.livro.autor_sobrenome = (
                registro.livro.autor.split()[-1] if registro.livro.autor else ''
            )
            registro.livro.cover_from = '#1e5aa8'
            registro.livro.cover_to   = '#0b1526'
            livros_lidos.append(registro)

    return render(request, 'biblioteca/acervo.html', {
        'aba':          'lidos',
        'livros_lidos': livros_lidos,
        'lidos_count':  lidos_count,
        **_acervo_counts(request.user),
    })

@login_required
def reservados(request):
    reservas_ativas = Reserva.objects.filter(
        usuario=request.user,
        status__in=['pendente', 'fila', 'aceita', 'aguardando_retirada'],
    ).select_related('livro').order_by('-data_reserva')

    for r in reservas_ativas:
        r.livro.cover_from = '#1e5aa8'
        r.livro.cover_to   = '#0b1526'
        r.livro.autor_sobrenome = r.livro.autor.split()[-1] if r.livro.autor else ''

    return render(request, 'biblioteca/acervo.html', {
        'aba': 'reservados',
        'reservas': reservas_ativas,
        **_acervo_counts(request.user),
    })

# ──────────────────────────────────────────
# LISTAS
# ──────────────────────────────────────────

@login_required
def detalhe_lista(request, lista_id):
    lista = get_object_or_404(Lista, id=lista_id, usuario=request.user)
    return render(request, 'biblioteca/detalhe_lista.html', {'lista': lista})


@require_POST
def criar_lista(request):
    if not request.user.is_authenticated:
        return JsonResponse({'erro': 'Não autenticado'}, status=401)

    data = json.loads(request.body)
    nome = data.get('nome', '').strip()
    if not nome:
        return JsonResponse({'erro': 'Nome não pode ser vazio.'}, status=400)

    nova_lista = Lista.objects.create(usuario=request.user, nome=nome)
    return JsonResponse({'id': nova_lista.id, 'nome': nova_lista.nome, 'qtd_livros': 0})


@require_POST
def excluir_lista(request, lista_id):
    if not request.user.is_authenticated:
        return JsonResponse({'erro': 'Não autenticado'}, status=401)

    lista = get_object_or_404(Lista, id=lista_id, usuario=request.user)
    lista.delete()
    return JsonResponse({'status': 'ok'})


@require_POST
def renomear_lista(request, lista_id):
    if not request.user.is_authenticated:
        return JsonResponse({'erro': 'Não autenticado'}, status=401)

    lista = get_object_or_404(Lista, id=lista_id, usuario=request.user)
    data = json.loads(request.body)
    novo_nome = data.get('nome', '').strip()
    descricao = data.get('descricao', '').strip()

    if not novo_nome:
        return JsonResponse({'erro': 'Nome não pode ser vazio.'}, status=400)

    lista.nome = novo_nome
    lista.descricao = descricao if descricao else None
    lista.save()
    return JsonResponse({'id': lista.id, 'nome': lista.nome, 'descricao': lista.descricao})


@require_POST
def adicionar_livro_lista(request, lista_id):
    if not request.user.is_authenticated:
        return JsonResponse({'erro': 'Não autenticado'}, status=401)

    lista = get_object_or_404(Lista, id=lista_id, usuario=request.user)
    try:
        data = json.loads(request.body)
        livro_id = data.get('livro_id')
        livro = Livro.objects.filter(id_livro=livro_id).first()
        if not livro:
            livro = Livro.objects.filter(pk=livro_id).first()
        if not livro:
            return JsonResponse({'erro': f'Livro {livro_id} não encontrado.'}, status=404)
        lista.livros.add(livro)
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'erro': str(e)}, status=500)


@require_POST
def remover_livro_lista(request, lista_id):
    if not request.user.is_authenticated:
        return JsonResponse({'erro': 'Não autenticado'}, status=401)

    try:
        lista = Lista.objects.get(pk=lista_id)
    except Lista.DoesNotExist:
        return JsonResponse({'erro': 'Lista não encontrada.'}, status=404)

    if lista.usuario_id != request.user.pk:
        return JsonResponse({'erro': 'Sem permissão.'}, status=403)

    try:
        data = json.loads(request.body)
        livro_id = data.get('livro_id')
        livro = Livro.objects.filter(id_livro=livro_id).first()
        if not livro:
            livro = Livro.objects.filter(pk=livro_id).first()
        if not livro:
            return JsonResponse({'erro': f'Livro {livro_id} não encontrado.'}, status=404)
        lista.livros.remove(livro)
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'erro': str(e)}, status=500)


def listas_do_usuario(request):
    if not request.user.is_authenticated:
        return JsonResponse({'listas': []})

    livro_id = request.GET.get('livro_id')
    listas = Lista.objects.filter(usuario=request.user).prefetch_related('livros')
    resultado = []
    for l in listas:
        tem_livro = False
        if livro_id:
            tem_livro = l.livros.filter(id_livro=livro_id).exists()
        resultado.append({
            'id':         l.id,
            'nome':       l.nome,
            'qtd_livros': l.livros.count(),
            'tem_livro':  tem_livro,
        })
    return JsonResponse({'listas': resultado})


@require_POST
def deletar_lista(request, pk):
    lista = get_object_or_404(Lista, pk=pk)
    lista.delete()
    return JsonResponse({'sucesso': 'Lista deletada.'})


# ──────────────────────────────────────────
# LIVROS LIDOS
# ──────────────────────────────────────────

@require_POST
@login_required
def marcar_livro_lido(request, livro_id):
    livro = get_object_or_404(Livro, id_livro=livro_id)
    _, criado = LivroLido.objects.get_or_create(usuario=request.user, livro=livro)
    return JsonResponse({'status': 'ok', 'criado': criado})


@require_POST
@login_required
def desmarcar_livro_lido(request, livro_id):
    livro = get_object_or_404(Livro, id_livro=livro_id)
    LivroLido.objects.filter(usuario=request.user, livro=livro).delete()
    return JsonResponse({'status': 'ok'})


# ──────────────────────────────────────────
# CONFIGURAÇÕES DO USUÁRIO
# ──────────────────────────────────────────

@require_POST
@login_required
def salvar_perfil(request):
    data = json.loads(request.body)
    apelido = data.get('apelido', '').strip()

    if apelido:
        if Usuario.objects.filter(apelido=apelido).exclude(pk=request.user.pk).exists():
            return JsonResponse({'erro': 'Esse apelido já está em uso.'}, status=400)
        request.user.apelido = apelido
    else:
        request.user.apelido = None

    request.user.save()
    return JsonResponse({'sucesso': True})


@require_POST
@login_required
def alterar_senha(request):
    data = json.loads(request.body)
    senha_atual = data.get('senha_atual', '').strip()
    senha_nova = data.get('senha_nova', '').strip()

    if not request.user.check_password(senha_atual):
        return JsonResponse({'erro': 'Senha atual incorreta.'}, status=400)
    if len(senha_nova) < 8:
        return JsonResponse({'erro': 'A nova senha deve ter pelo menos 8 caracteres.'}, status=400)

    request.user.set_password(senha_nova)
    request.user.save()

    from django.contrib.auth import update_session_auth_hash
    update_session_auth_hash(request, request.user)

    return JsonResponse({'sucesso': True})


@require_POST
@login_required
def salvar_notif(request):
    data = json.loads(request.body)
    # Por enquanto apenas confirma — lógica de notificação real vem depois
    return JsonResponse({'sucesso': True})

# ──────────────────────────────────────────
# NOTIFICAÇÕES
# ──────────────────────────────────────────

@login_required
def notificacoes_listar(request):
    """Lista as últimas 50 notificações do usuário logado, mais recentes primeiro."""
    notifs = Notificacao.objects.filter(destinatario=request.user).order_by('-criada_em')[:50]

    resultado = [{
        'id':           n.pk,
        'tipo':         n.tipo,
        'titulo':       n.titulo,
        'mensagem':     n.mensagem,
        'lida':         n.lida,
        'requer_acao':  n.requer_acao,
        'acao_tomada':  n.acao_tomada,
        'criada_em':    n.criada_em.strftime('%d/%m/%Y %H:%M'),
        'reserva_id':   n.reserva_id,
    } for n in notifs]

    nao_lidas = Notificacao.objects.filter(destinatario=request.user, lida=False).count()

    return JsonResponse({'notificacoes': resultado, 'nao_lidas': nao_lidas})


@login_required
def notificacoes_nao_lidas(request):
    """Endpoint leve, só pro sininho atualizar o número do badge periodicamente."""
    total = Notificacao.objects.filter(destinatario=request.user, lida=False).count()
    return JsonResponse({'total': total})


@require_POST
@login_required
def notificacoes_marcar_lida(request, pk):
    notif = get_object_or_404(Notificacao, pk=pk, destinatario=request.user)
    notif.lida = True
    notif.save(update_fields=['lida'])
    return JsonResponse({'sucesso': True})


@require_POST
@login_required
def notificacoes_marcar_todas_lidas(request):
    Notificacao.objects.filter(destinatario=request.user, lida=False).update(lida=True)
    return JsonResponse({'sucesso': True})


@require_POST
@login_required
def notificacoes_responder(request, pk):
    notif = get_object_or_404(Notificacao, pk=pk, destinatario=request.user)
    resposta = json.loads(request.body).get('resposta', '').strip()

    if resposta not in ('aceita', 'recusada'):
        return JsonResponse({'erro': 'Resposta inválida.'}, status=400)
    if not notif.requer_acao:
        return JsonResponse({'erro': 'Esta notificação não requer ação.'}, status=400)
    if notif.acao_tomada:
        return JsonResponse({'erro': 'Esta notificação já foi respondida.'}, status=400)

    if notif.reserva_id:
        try:
            if notif.tipo == 'reserva_pendente':
                aceitar_reserva(notif.reserva) if resposta == 'aceita' else recusar_reserva(notif.reserva)
            elif notif.tipo == 'devolucao_com_reserva' and resposta == 'aceita':
                confirmar_disponibilidade(notif.reserva)
        except ReservaError as e:
            return JsonResponse({'erro': str(e)}, status=400)

    notif.acao_tomada = resposta
    notif.lida = True
    notif.save(update_fields=['acao_tomada', 'lida'])

    return JsonResponse({'sucesso': True, 'acao_tomada': notif.acao_tomada})

@require_POST
@login_required
def notificacoes_excluir(request, pk):
    notif = get_object_or_404(Notificacao, pk=pk, destinatario=request.user)
    notif.delete()
    return JsonResponse({'sucesso': True})


@require_POST
@login_required
def notificacoes_excluir_todas(request):
    Notificacao.objects.filter(destinatario=request.user).delete()
    return JsonResponse({'sucesso': True})
# ──────────────────────────────────────────
# ADMIN / BIBLIOTECÁRIA
# ──────────────────────────────────────────

def cadastrar_livro(request):
    if request.method == 'GET':
        return render(request, 'biblioteca/cad-livro.html')

    data = json.loads(request.body)
    titulo      = data.get('titulo', '').strip()
    autor       = data.get('autor', '').strip()
    editora     = data.get('editora', '').strip()
    categoria   = data.get('categoria', '').strip()
    colecao     = data.get('colecao', '').strip()
    prateleira  = data.get('prateleira', '').strip()
    codigo_base = data.get('codigo_base', '').strip()
    quantidade  = int(data.get('quantidade', 1))
    observacoes = data.get('observacoes', '').strip()

    if not titulo or not prateleira or not quantidade:
        return JsonResponse({'erro': 'Preencha todos os campos obrigatórios.'}, status=400)

    livro = Livro.objects.create(
        titulo=titulo,
        autor=autor or None,
        editora=editora or None,
        categoria=categoria or None,
        colecao=colecao or None,
        prateleira=prateleira,
        codigo_base=codigo_base or None,
        quantidade=quantidade,
        observacoes=observacoes or None,
    )

    for i in range(1, quantidade + 1):
        codigo_completo = f"{codigo_base}/V{i}" if codigo_base else f"{livro.id_livro}/V{i}"
        Exemplar.objects.create(
            livro=livro,
            codigo_variante=f"V{i}",
            codigo_completo=codigo_completo,
            status='disponivel',
        )

    return JsonResponse({
        'sucesso':            True,
        'id':                 livro.id_livro,
        'titulo':             livro.titulo,
        'exemplares_criados': quantidade,
    })


@require_POST
@staff_member_required
def editar_livro(request, livro_id):
    livro = get_object_or_404(Livro, id_livro=livro_id)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erro': 'JSON inválido.'}, status=400)

    campos_texto = [
        'titulo', 'autor', 'editora', 'isbn', 'codigo_base',
        'categoria', 'colecao', 'prateleira', 'capa_url', 'sinopse',
    ]
    for campo in campos_texto:
        if campo in data:
            valor = data[campo].strip()
            setattr(livro, campo, valor if valor else None)

    if 'quantidade' in data:
        try:
            livro.quantidade = max(0, int(data['quantidade']))
        except (ValueError, TypeError):
            return JsonResponse({'erro': 'Quantidade inválida.'}, status=400)

    livro.save()
    return JsonResponse({'sucesso': True})


@require_POST
@staff_member_required
def excluir_livro(request, id_livro):
    livro = get_object_or_404(Livro, id_livro=id_livro)
    livro.delete()
    return JsonResponse({'sucesso': True})


def emprestimos(request):
    hoje = date.today()

    CORES = ['#1e5aa8', '#7c3aed', '#059669', '#b45309', '#dc2626', '#0891b2']

    emprestimos_ativos = Emprestimo.objects.filter(
        data_devolucao_real__isnull=True
    ).select_related('exemplar__livro', 'usuario').order_by('data_devolucao_prevista')

    vencem_hoje     = emprestimos_ativos.filter(data_devolucao_prevista=hoje).count()
    atrasados       = sum(1 for e in emprestimos_ativos if e.esta_atrasado())
    devolvidos_hoje = Emprestimo.objects.filter(data_devolucao_real=hoje).count()

    def enrich(emp, hoje, CORES):
        delta = (emp.data_devolucao_prevista - hoje).days
        emp.vence_hoje     = (delta == 0)
        emp.dias_restantes = max(delta, 0)
        emp.dias_atraso    = abs(delta) if delta < 0 else 0
        partes = emp.usuario.get_full_name().split() if emp.usuario else []
        if len(partes) >= 2:
            emp.iniciais_av = partes[0][0].upper() + partes[-1][0].upper()
        elif partes:
            emp.iniciais_av = partes[0][0].upper()
        else:
            emp.iniciais_av = '?'
        emp.cor_avatar           = CORES[emp.pk % len(CORES)]
        emp.nome_aluno           = emp.usuario.get_full_name() if emp.usuario else 'Desconhecido'
        emp.turma                = emp.usuario.serie if emp.usuario else ''
        emp.matricula_aluno      = emp.usuario.matricula if emp.usuario else ''
        emp.codigo_catalografico = emp.exemplar.codigo_completo if emp.exemplar else ''
        emp.livro_obj            = emp.exemplar.livro if emp.exemplar else None
        return emp

    emprestimos_enriched = [enrich(e, hoje, CORES) for e in emprestimos_ativos]

    historico_qs = Emprestimo.objects.filter(
        data_devolucao_real__isnull=False
    ).select_related('exemplar__livro', 'usuario').order_by('-data_devolucao_real')[:100]

    historico_enriched = []
    for emp in historico_qs:
        delta = (emp.data_devolucao_prevista - emp.data_devolucao_real).days
        emp.vence_hoje     = False
        emp.dias_restantes = 0
        emp.dias_atraso    = 0
        emp.foi_devolvido_atrasado = emp.data_devolucao_real > emp.data_devolucao_prevista
        partes = emp.usuario.get_full_name().split() if emp.usuario else []
        if len(partes) >= 2:
            emp.iniciais_av = partes[0][0].upper() + partes[-1][0].upper()
        elif partes:
            emp.iniciais_av = partes[0][0].upper()
        else:
            emp.iniciais_av = '?'
        emp.cor_avatar           = CORES[emp.pk % len(CORES)]
        emp.nome_aluno           = emp.usuario.get_full_name() if emp.usuario else 'Desconhecido'
        emp.turma                = emp.usuario.serie if emp.usuario else ''
        emp.matricula_aluno      = emp.usuario.matricula if emp.usuario else ''
        emp.codigo_catalografico = emp.exemplar.codigo_completo if emp.exemplar else ''
        emp.livro_obj            = emp.exemplar.livro if emp.exemplar else None
        historico_enriched.append(emp)

    return render(request, 'biblioteca/emp-livro.html', {
        'emprestimos':     emprestimos_enriched,
        'historico':       historico_enriched,
        'vencem_hoje':     vencem_hoje,
        'atrasados':       atrasados,
        'devolvidos_hoje': devolvidos_hoje,
    })


@require_POST
def criar_emprestimo(request):
    data = json.loads(request.body)
    codigo_completo = data.get('codigo_completo', '').strip()
    nome_aluno      = data.get('nome_aluno', '').strip()
    turma           = data.get('turma', '').strip()
    usuario_id      = data.get('usuario_id')
    data_emp        = data.get('data_emprestimo', '').strip()
    data_dev        = data.get('data_devolucao_prevista', '').strip()

    if not all([codigo_completo, nome_aluno, data_emp]):
        return JsonResponse({'erro': 'Preencha todos os campos obrigatórios.'}, status=400)

    exemplar = Exemplar.objects.filter(codigo_completo=codigo_completo).first()
    if not exemplar:
        return JsonResponse({'erro': 'Exemplar não encontrado.'}, status=404)
    if exemplar.status != 'disponivel':
        return JsonResponse({'erro': 'Exemplar não está disponível.'}, status=400)

    usuario = None
    if usuario_id:
        usuario = Usuario.objects.filter(pk=usuario_id).first()

    if not usuario:
        partes = nome_aluno.split()
        if partes:
            q = Q(first_name__icontains=partes[0])
            if len(partes) > 1:
                q &= Q(last_name__icontains=partes[-1])
            usuario = Usuario.objects.filter(q).first()

    if not usuario:
        return JsonResponse({'erro': f'Aluno "{nome_aluno}" não encontrado. Verifique o nome.'}, status=404)

    emprestimos_ativos = Emprestimo.objects.filter(
        usuario=usuario,
        data_devolucao_real__isnull=True
    ).count()
    if emprestimos_ativos >= 2:
        return JsonResponse({
            'erro': f'{usuario.get_full_name()} já possui 2 empréstimos ativos (limite máximo).'
        }, status=400)

    data_emprestimo = date.fromisoformat(data_emp)
    data_devolucao  = date.fromisoformat(data_dev) if data_dev else data_emprestimo + timezone.timedelta(days=15)

    emp = Emprestimo.objects.create(
        exemplar=exemplar,
        usuario=usuario,
        data_emprestimo=data_emprestimo,
        data_devolucao_prevista=data_devolucao,
        observacoes=data.get('observacoes', '')
    )

    exemplar.status = 'emprestado'
    exemplar.save()

    return JsonResponse({
        'id':                      emp.pk,
        'titulo':                  exemplar.livro.titulo,
        'codigo_completo':         exemplar.codigo_completo,
        'usuario':                 usuario.get_full_name(),
        'turma':                   usuario.serie or turma,
        'data_emprestimo':         str(emp.data_emprestimo),
        'data_devolucao_prevista': str(emp.data_devolucao_prevista),
        'atrasado':                emp.esta_atrasado(),
        'observacoes':             emp.observacoes,
    })


@require_POST
def devolver_emprestimo(request, pk):
    emprestimo = get_object_or_404(Emprestimo, pk=pk)
    emprestimo.data_devolucao_real = timezone.now().date()
    emprestimo.save()

    tem_reserva_aceita = Reserva.objects.filter(
        livro=emprestimo.exemplar.livro, status='aceita'
    ).exists()

    emprestimo.exemplar.status = 'reservado' if tem_reserva_aceita else 'disponivel'
    emprestimo.exemplar.save()

    if tem_reserva_aceita:
        verificar_devolucao_com_reserva(emprestimo)

    return JsonResponse({'sucesso': 'Livro devolvido com sucesso.'})


@require_POST
def renovar_emprestimo(request, pk):
    emprestimo = get_object_or_404(Emprestimo, pk=pk)
    hoje = timezone.now().date()
    if emprestimo.data_devolucao_prevista < hoje:
        return JsonResponse({'erro': 'Não é possível renovar um empréstimo atrasado.', 'status': 'Erro'}, status=400)
    emprestimo.data_devolucao_prevista += timezone.timedelta(days=15)
    emprestimo.save()
    return JsonResponse({
        'sucesso':   f'Renovado com sucesso. Nova data: {emprestimo.data_devolucao_prevista}',
        'nova_data': str(emprestimo.data_devolucao_prevista),
        'status':    'Sucesso'
    })


@require_POST
def excluir_historico(request, pk):
    emprestimo = get_object_or_404(Emprestimo, pk=pk)
    if emprestimo.data_devolucao_real is None:
        return JsonResponse({'erro': 'Não é possível excluir um empréstimo ainda ativo.'}, status=400)
    emprestimo.delete()
    return JsonResponse({'sucesso': 'Registro excluído do histórico.'})


@require_POST
@login_required
def cancelar_reserva(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    try:
        cancelar_reserva_usuario(reserva, request.user)
    except ReservaError as e:
        return JsonResponse({'erro': str(e)}, status=403)
    return JsonResponse({'sucesso': 'Reserva cancelada com sucesso.'})  


# ──────────────────────────────────────────
# ALUNOS
# ──────────────────────────────────────────

def alunos(request):
    CORES = ['#1e5aa8', '#7c3aed', '#059669', '#b45309', '#dc2626', '#0891b2', '#be185d']

    alunos_qs = Usuario.objects.filter(
        tipo_usuario='aluno'
    ).order_by('serie', 'first_name', 'last_name')

    for aluno in alunos_qs:
        partes = aluno.get_full_name().split()
        aluno.iniciais   = (partes[0][0] + partes[-1][0]).upper() if len(partes) >= 2 else partes[0][0].upper() if partes else '?'
        aluno.cor_avatar = CORES[aluno.pk % len(CORES)]

    alunos_por_turma = defaultdict(list)
    for aluno in alunos_qs:
        alunos_por_turma[aluno.serie or 'Sem turma'].append(aluno)

    turmas = sorted(alunos_por_turma.keys())

    return render(request, 'biblioteca/alunos.html', {
        'alunos_por_turma': dict(sorted(alunos_por_turma.items())),
        'turmas':           turmas,
        'total_alunos':     alunos_qs.count(),
        'total_turmas':     len(turmas),
    })

@require_POST
@staff_member_required
def excluir_aluno(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if usuario.tipo_usuario == 'bibliotecario':
        return JsonResponse({'erro': 'Não é possível excluir um bibliotecário.'}, status=403)
    usuario.delete()
    return JsonResponse({'sucesso': True})

@require_POST
def importar_alunos(request):
    data = json.loads(request.body)
    alunos = data.get('alunos', [])
    criados = atualizados = 0

    for item in alunos:
        nome      = item.get('nome', '').strip()
        matricula = item.get('matricula', '').strip()
        turma     = item.get('turma', '').strip()

        if not nome or not matricula:
            continue

        partes     = nome.split()
        first_name = partes[0] if partes else ''
        last_name  = ' '.join(partes[1:]) if len(partes) > 1 else ''

        usuario, criado = Usuario.objects.update_or_create(
            matricula=matricula,
            defaults={
                'first_name':      first_name,
                'last_name':       last_name,
                'serie':           turma,
                'tipo_usuario':    'aluno',
                'username':        matricula,
                'primeiro_acesso': True,
            }
        )
        if criado:
            usuario.set_unusable_password()
            usuario.save()
            criados += 1
        else:
            atualizados += 1

    return JsonResponse({'criados': criados, 'atualizados': atualizados})

@require_POST
@staff_member_required
def limpar_alunos(request):
    data = json.loads(request.body)
    escopo = data.get('escopo', '').strip()

    if escopo == 'todos':
        removidos, _ = Usuario.objects.filter(tipo_usuario='aluno').delete()

    elif escopo == 'turmas':
        turmas = data.get('turmas', [])
        if not turmas:
            return JsonResponse({'erro': 'Nenhuma turma informada.'}, status=400)
        removidos, _ = Usuario.objects.filter(
            tipo_usuario='aluno',
            serie__in=turmas
        ).delete()

    else:
        return JsonResponse({'erro': 'Escopo inválido.'}, status=400)

    return JsonResponse({'removidos': removidos})

# ──────────────────────────────────────────
# BUSCA
# ──────────────────────────────────────────

def buscar_livro(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'livros': []})

    livros = Livro.objects.filter(titulo__icontains=q)[:8]
    resultado = []
    for livro in livros:
        exemplares_list = list(
            Exemplar.objects.filter(livro=livro, status='disponivel')
            .values('id_exemplar', 'codigo_completo', 'codigo_variante')
        )
        resultado.append({
            'id':                     livro.id_livro,
            'titulo':                 livro.titulo,
            'autor':                  livro.autor or '',
            'prateleira':             livro.prateleira or '',
            'capa_url':               livro.capa_url or '',
            'total_exemplares':       livro.quantidade,
            'exemplares_disponiveis': exemplares_list,
            'qtd_disponivel':         len(exemplares_list),
        })

    return JsonResponse({'livros': resultado})


def buscar_usuario(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'usuarios': []})

    usuarios = Usuario.objects.filter(
        Q(first_name__icontains=q) | Q(last_name__icontains=q)
    ).exclude(tipo_usuario='bibliotecario')[:8]

    resultado = []
    for u in usuarios:
        emprestimos_ativos = Emprestimo.objects.filter(
            usuario=u, data_devolucao_real__isnull=True
        ).count()
        resultado.append({
            'id':                 u.pk,
            'nome':               u.get_full_name() or u.username,
            'matricula':          str(u.matricula) if u.matricula else '',
            'serie':              u.serie or '',
            'tipo':               u.tipo_usuario,
            'emprestimos_ativos': emprestimos_ativos,
            'pode_emprestar':     emprestimos_ativos < 2,
        })

    return JsonResponse({'usuarios': resultado})


# ── Estilos compartilhados ─────────────────

def _borda():
    thin = Side(style='thin', color='BBCCE0')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

AZUL_ESCURO = '1E3A5F'
AZUL_HEADER = '1E5AA8'
BRANCO      = 'FFFFFF'
CINZA       = 'F0F4F9'
AMARELO     = 'FFF3CD'
VERMELHO    = 'C0392B'


# ── Página principal de exportação ────────

@staff_member_required
def exportar(request):
    total_livros     = Livro.objects.count()
    total_emprestimos = Emprestimo.objects.count()
    total_alunos     = Usuario.objects.filter(tipo_usuario='aluno').count()
    return render(request, 'biblioteca/exportar.html', {
        'total_livros':      total_livros,
        'total_emprestimos': total_emprestimos,
        'total_alunos':      total_alunos,
    })


# ── Exportar Acervo de Livros ──────────────

@staff_member_required
def exportar_acervo(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Acervo'
    borda = _borda()

    # Linha 1 vazia
    ws.append([None])

    # Linha 2 — título
    ws.merge_cells('A2:I2')
    escola = 'ACERVO DE LIVROS - EREM DR. JAIME MONTEIRO'
    ws['A2'] = escola
    ws['A2'].font      = Font(name='Arial', bold=True, size=13, color=BRANCO)
    ws['A2'].fill      = PatternFill('solid', fgColor=AZUL_ESCURO)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 30

    # Linha 3 vazia
    ws.append([None])

    # Linha 4 — cabeçalho
    headers = ['Nº DE REGISTRO', 'TÍTULO', 'AUTOR(A)', 'EDITORA',
               'CATEGORIA', 'COLEÇÃO', 'QTD.', 'CATÁLOGO', 'OBSERVAÇÕES']
    ws.append(headers)
    for col in range(1, 10):
        c = ws.cell(row=4, column=col)
        c.font      = Font(name='Arial', bold=True, size=10, color=BRANCO)
        c.fill      = PatternFill('solid', fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borda
    ws.row_dimensions[4].height = 20

    # Dados
    livros = Livro.objects.all().order_by('id_livro')
    for i, livro in enumerate(livros):
        row_num = 5 + i
        ws.append([
            i + 1,
            livro.titulo or '',
            livro.autor or '',
            livro.editora or '',
            livro.categoria or '',
            livro.colecao or '',
            livro.quantidade or 0,
            livro.codigo_base or '',
            livro.observacoes or '',
        ])
        fill = CINZA if i % 2 == 0 else BRANCO
        for col in range(1, 10):
            c = ws.cell(row=row_num, column=col)
            c.font      = Font(name='Arial', size=10)
            c.fill      = PatternFill('solid', fgColor=fill)
            c.alignment = Alignment(vertical='center', wrap_text=(col == 2))
            c.border    = borda
        ws.row_dimensions[row_num].height = 18

    # Linha de total
    total_row = 5 + len(livros)
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=1).font = Font(name='Arial', bold=True, size=10, color=BRANCO)
    ws.cell(row=total_row, column=1).fill = PatternFill('solid', fgColor=AZUL_ESCURO)
    ws.cell(row=total_row, column=7, value=f'=SUM(G5:G{total_row - 1})')
    ws.cell(row=total_row, column=7).font = Font(name='Arial', bold=True, size=10, color=BRANCO)
    ws.cell(row=total_row, column=7).fill = PatternFill('solid', fgColor=AZUL_ESCURO)

    # Larguras
    for i, w in enumerate([14, 48, 30, 24, 18, 22, 7, 22, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Congelar cabeçalho
    ws.freeze_panes = 'A5'

    return _resposta_xlsx(wb, f'Acervo_Biblioteca_{date.today().year}.xlsx')


# ── Exportar Empréstimos ───────────────────

@staff_member_required
def exportar_emprestimos(request):
    ano   = request.GET.get('ano', date.today().year)
    todos = request.GET.get('todos', '0') == '1'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Empréstimos'
    borda = _borda()

    # Linha 1 — título + prazo padrão
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Planilha de retirada de livros da biblioteca - EREM DR. JAIME MONTEIRO'
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color=BRANCO)
    ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_ESCURO)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('G1:H1')
    ws['G1'] = 'DIAS ATÉ A DEVOLUÇÃO:'
    ws['G1'].font      = Font(name='Arial', bold=True, size=10)
    ws['G1'].alignment = Alignment(horizontal='right', vertical='center')
    ws['I1'] = 15
    ws['I1'].font      = Font(name='Arial', bold=True, size=11, color=VERMELHO)
    ws['I1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Linha 2 — cabeçalho
    headers = ['Em atraso', 'ALUNO', 'TURMA', 'TELEFONE DE CONTATO',
               'TÍTULO DE LIVRO', 'DATA DO EMPRÉSTIMO', 'DATA DA DEVOLUÇÃO', 'DIAS', 'OBSERVAÇÃO']
    ws.append(headers)
    for col in range(1, 10):
        c = ws.cell(row=2, column=col)
        c.font      = Font(name='Arial', bold=True, size=10, color=BRANCO)
        c.fill      = PatternFill('solid', fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borda
    ws.row_dimensions[2].height = 20

    # Busca empréstimos
    qs = Emprestimo.objects.select_related(
        'usuario', 'exemplar__livro'
    ).order_by('usuario__first_name', 'data_emprestimo')

    if not todos:
        qs = qs.filter(data_emprestimo__year=ano)

    hoje = date.today()
    for i, emp in enumerate(qs):
        row_num = 3 + i
        atrasado = emp.esta_atrasado()
        delta = (emp.data_devolucao_prevista - (emp.data_devolucao_real or hoje)).days

        ws.append([
            'SIM' if atrasado else 'NÃO',
            emp.usuario.get_full_name() if emp.usuario else '',
            emp.usuario.serie if emp.usuario else '',
            emp.usuario.telefone if (emp.usuario and hasattr(emp.usuario, 'telefone')) else '',
            emp.exemplar.livro.titulo if emp.exemplar and emp.exemplar.livro else '',
            emp.data_emprestimo.strftime('%d/%m/%Y') if emp.data_emprestimo else '',
            emp.data_devolucao_prevista.strftime('%d/%m/%Y') if emp.data_devolucao_prevista else '',
            delta if not emp.data_devolucao_real else 0,
            emp.observacoes or '',
        ])

        fill = AMARELO if atrasado else (CINZA if i % 2 == 0 else BRANCO)
        for col in range(1, 10):
            c = ws.cell(row=row_num, column=col)
            cor_fonte = VERMELHO if (atrasado and col == 1) else '000000'
            c.font      = Font(name='Arial', size=10, bold=(col == 1 and atrasado), color=cor_fonte)
            c.fill      = PatternFill('solid', fgColor=fill)
            c.alignment = Alignment(vertical='center')
            c.border    = borda
        ws.row_dimensions[row_num].height = 18

    # Larguras
    for i, w in enumerate([10, 38, 10, 22, 38, 18, 18, 8, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    sufixo = 'Completo' if todos else str(ano)
    return _resposta_xlsx(wb, f'Emprestimos_Biblioteca_{sufixo}.xlsx')


# ── Exportar Lista de Alunos ───────────────

@staff_member_required
def exportar_alunos(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Alunos'
    borda = _borda()

    # Linha 1 — título
    ws.merge_cells('A1:E1')
    ws['A1'] = f'LISTA DE ALUNOS - EREM DR. JAIME MONTEIRO — {date.today().year}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color=BRANCO)
    ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_ESCURO)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.append([None])  # Linha 2 vazia

    # Linha 3 — cabeçalho
    headers = ['Nº', 'NOME COMPLETO', 'MATRÍCULA', 'TURMA', 'E-MAIL']
    ws.append(headers)
    for col in range(1, 6):
        c = ws.cell(row=3, column=col)
        c.font      = Font(name='Arial', bold=True, size=10, color=BRANCO)
        c.fill      = PatternFill('solid', fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borda
    ws.row_dimensions[3].height = 20

    alunos = Usuario.objects.filter(
        tipo_usuario='aluno'
    ).order_by('serie', 'first_name', 'last_name')

    turma_atual = None
    num = 0
    row_num = 4

    for aluno in alunos:
        # Linha separadora de turma
        if aluno.serie != turma_atual:
            turma_atual = aluno.serie or 'Sem turma'
            ws.merge_cells(f'A{row_num}:E{row_num}')
            ws.cell(row=row_num, column=1, value=f'  {turma_atual}')
            ws.cell(row=row_num, column=1).font      = Font(name='Arial', bold=True, size=10, color=BRANCO)
            ws.cell(row=row_num, column=1).fill      = PatternFill('solid', fgColor='2C5282')
            ws.cell(row=row_num, column=1).alignment = Alignment(vertical='center')
            ws.row_dimensions[row_num].height = 18
            row_num += 1
            num = 0

        num += 1
        fill = CINZA if num % 2 == 0 else BRANCO
        dados = [num, aluno.get_full_name(), str(aluno.matricula or ''),
                 aluno.serie or '', aluno.email or '']

        for col, val in enumerate(dados, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.font      = Font(name='Arial', size=10)
            c.fill      = PatternFill('solid', fgColor=fill)
            c.alignment = Alignment(vertical='center',
                                    horizontal='center' if col in (1, 3, 4) else 'left')
            c.border    = borda
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    for i, w in enumerate([5, 42, 14, 10, 32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'

    return _resposta_xlsx(wb, f'Alunos_Biblioteca_{date.today().year}.xlsx')


# ── Helper: resposta HTTP com xlsx ─────────

def _resposta_xlsx(wb: Workbook, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
# ──────────────────────────────────────────
# API REST
# ──────────────────────────────────────────

class LivroViewset(viewsets.ModelViewSet):
    queryset = Livro.objects.all()
    serializer_class = LivroSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['titulo', 'autor']

# ──────────────────────────────────────────
# TURMAS (banco de dados)
# ──────────────────────────────────────────

def listar_turmas(request):
    """Retorna todas as turmas cadastradas no banco."""
    turmas = Turma.objects.values_list('nome', flat=True)
    return JsonResponse({'turmas': list(turmas)})


@require_POST
@staff_member_required
def salvar_turmas(request):
    """
    Recebe lista de turmas e substitui tudo no banco.
    Payload: { "turmas": ["1° A", "2° B", ...] }
    """
    data  = json.loads(request.body)
    nomes = [t.strip() for t in data.get('turmas', []) if t.strip()]

    # Remove todas e recria na ordem recebida
    Turma.objects.all().delete()
    for i, nome in enumerate(nomes):
        Turma.objects.create(nome=nome, ordem=i)

    return JsonResponse({'sucesso': True, 'total': len(nomes)})


# ──────────────────────────────────────────
# CARGOS
# ──────────────────────────────────────────

def listar_cargos(request):
    """Retorna todos os cargos com suas abas."""
    cargos = list(Cargo.objects.values('id', 'nome', 'abas'))
    return JsonResponse({'cargos': cargos})


@require_POST
@staff_member_required
def salvar_cargo(request):
    """
    Cria ou atualiza um cargo.
    Payload: { "id": 1 (opcional), "nome": "Monitor", "abas": ["emprestimos"] }
    """
    data  = json.loads(request.body)
    cargo_id = data.get('id')
    nome  = data.get('nome', '').strip()
    abas  = data.get('abas', [])

    if not nome:
        return JsonResponse({'erro': 'Nome do cargo é obrigatório.'}, status=400)

    # Valida abas recebidas
    slugs_validos = [slug for slug, _ in ABAS_DISPONIVEIS]
    abas_validas  = [a for a in abas if a in slugs_validos]

    if cargo_id:
        cargo = get_object_or_404(Cargo, pk=cargo_id)
        # Verifica duplicidade de nome em outro cargo
        if Cargo.objects.filter(nome=nome).exclude(pk=cargo_id).exists():
            return JsonResponse({'erro': 'Já existe um cargo com esse nome.'}, status=400)
        cargo.nome = nome
        cargo.abas = abas_validas
        cargo.save()
        criado = False
    else:
        if Cargo.objects.filter(nome=nome).exists():
            return JsonResponse({'erro': 'Já existe um cargo com esse nome.'}, status=400)
        cargo  = Cargo.objects.create(nome=nome, abas=abas_validas)
        criado = True

    return JsonResponse({
        'sucesso': True,
        'criado':  criado,
        'cargo':   {'id': cargo.pk, 'nome': cargo.nome, 'abas': cargo.abas},
    })


@require_POST
@staff_member_required
def excluir_cargo(request, pk):
    """Remove um cargo. Usuários com esse cargo ficam sem cargo (SET_NULL)."""
    cargo = get_object_or_404(Cargo, pk=pk)
    cargo.delete()
    return JsonResponse({'sucesso': True})


@require_POST
@staff_member_required
def atribuir_cargo(request, pk):
    """
    Atribui ou remove um cargo de um usuário.
    Payload: { "cargo_id": 1 }  — use null para remover o cargo.
    """
    usuario  = get_object_or_404(Usuario, pk=pk)
    data     = json.loads(request.body)
    cargo_id = data.get('cargo_id')

    if cargo_id:
        cargo = get_object_or_404(Cargo, pk=cargo_id)
        usuario.cargo = cargo
    else:
        usuario.cargo = None

    usuario.save()
    return JsonResponse({
        'sucesso':    True,
        'cargo_nome': usuario.cargo.nome if usuario.cargo else None,
    })

MESES_PT = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

MESES_PT_ABREV = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                   'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']


def _variacao_pct(atual, anterior):
    """Retorna a variação percentual entre dois períodos."""
    if anterior == 0:
        return 100.0 if atual > 0 else 0.0
    return round(((atual - anterior) / anterior) * 100, 1)


def _emprestimo_esta_atrasado_em(emp, referencia):
    """
    Considera atrasado se:
    - já foi devolvido, mas depois do prazo; ou
    - ainda não foi devolvido e a data de referência já passou do prazo.
    """
    if emp.data_devolucao_real:
        return emp.data_devolucao_real > emp.data_devolucao_prevista
    return referencia > emp.data_devolucao_prevista


def _contar_atrasados(queryset, referencia):
    total = 0
    for emp in queryset.only('data_devolucao_real', 'data_devolucao_prevista'):
        if _emprestimo_esta_atrasado_em(emp, referencia):
            total += 1
    return total


def _dados_grafico(intervalo, turma_filtro):
    """
    Monta os dados do gráfico "Empréstimos por dia".
    - 30 dias  -> granularidade diária
    - 90 dias  -> granularidade semanal
    - 365 dias -> granularidade mensal
    """
    hoje = date.today()
    intervalo = str(intervalo)
    dias = {'30': 30, '90': 90, '365': 365}.get(intervalo, 30)
    data_inicio = hoje - timedelta(days=dias - 1)

    qs = Emprestimo.objects.filter(
        data_emprestimo__gte=data_inicio,
        data_emprestimo__lte=hoje,
    )
    if turma_filtro:
        qs = qs.filter(usuario__serie=turma_filtro)

    labels, valores, hoje_index = [], [], None

    if intervalo == '30':
        contagem = defaultdict(int)
        for row in qs.values('data_emprestimo').annotate(total=Count('id')):
            contagem[row['data_emprestimo']] = row['total']

        d, i = data_inicio, 0
        while d <= hoje:
            labels.append(d.strftime('%d/%m'))
            valores.append(contagem.get(d, 0))
            if d == hoje:
                hoje_index = i
            d += timedelta(days=1)
            i += 1

    elif intervalo == '90':
        agregados = (
            qs.annotate(semana=TruncWeek('data_emprestimo'))
              .values('semana')
              .annotate(total=Count('id'))
        )
        mapa = {row['semana']: row['total'] for row in agregados}

        semana_atual = hoje - timedelta(days=hoje.weekday())
        d = data_inicio - timedelta(days=data_inicio.weekday())
        i = 0
        while d <= semana_atual:
            labels.append(d.strftime('%d/%m'))
            valores.append(mapa.get(d, 0))
            if d == semana_atual:
                hoje_index = i
            d += timedelta(days=7)
            i += 1

    else:  # 365 dias
        agregados = (
            qs.annotate(mes_ref=TruncMonth('data_emprestimo'))
              .values('mes_ref')
              .annotate(total=Count('id'))
        )
        mapa = {row['mes_ref']: row['total'] for row in agregados}

        mes_atual = date(hoje.year, hoje.month, 1)
        d = date(data_inicio.year, data_inicio.month, 1)
        i = 0
        while d <= mes_atual:
            labels.append(f"{MESES_PT_ABREV[d.month]}/{str(d.year)[2:]}")
            valores.append(mapa.get(d, 0))
            if d == mes_atual:
                hoje_index = i
            d = date(d.year + 1, 1, 1) if d.month == 12 else date(d.year, d.month + 1, 1)
            i += 1

    return {'labels': labels, 'valores': valores, 'hoje_index': hoje_index}


# ──────────────────────────────────────────
# RANKING: paleta de cores (livros mais lidos / alunos que mais leram)
# ──────────────────────────────────────────

PALETA_RANKING = ['#4f94d6', '#22c55e', '#ec4899', '#64748b', '#f59e0b']


def _cor_ranking(indice):
    return PALETA_RANKING[indice % len(PALETA_RANKING)]


def _grafico_barras_pdf(intervalo, turma_filtro):
    """
    Mesma lógica de _dados_grafico, mas devolve a altura de cada barra em %
    (0-100), pronta para renderizar em CSS puro no PDF (sem JS/Chart.js,
    que o WeasyPrint não executa).
    """
    dados = _dados_grafico(intervalo, turma_filtro)
    valores = dados['valores']
    hoje_index = dados['hoje_index']
    maximo = max(valores) if valores else 0

    barras = [{
        'altura': int((v / maximo) * 100) if maximo else 0,
        'hoje': (i == hoje_index),
    } for i, v in enumerate(valores)]

    return {'barras': barras, 'maximo': maximo}


def _montar_contexto_relatorio(periodo_str, turma_filtro, intervalo):
    """
    Monta todo o contexto de dados do relatório para um período (mês/ano) e
    turma opcionais. Reutilizado tanto pela página normal quanto pela
    exportação em PDF, para garantir que os números batem exatamente.
    """
    hoje = date.today()

    try:
        ano, mes = map(int, periodo_str.split('-'))
    except (ValueError, AttributeError, TypeError):
        ano, mes = hoje.year, hoje.month
        periodo_str = hoje.strftime('%Y-%m')

    data_inicio = date(ano, mes, 1)
    data_fim = date(ano, mes, monthrange(ano, mes)[1])

    # Período anterior (mês anterior), para calcular variações
    if mes == 1:
        ano_ant, mes_ant = ano - 1, 12
    else:
        ano_ant, mes_ant = ano, mes - 1
    data_inicio_ant = date(ano_ant, mes_ant, 1)
    data_fim_ant = date(ano_ant, mes_ant, monthrange(ano_ant, mes_ant)[1])

    emprestimos_periodo = Emprestimo.objects.filter(
        data_emprestimo__gte=data_inicio, data_emprestimo__lte=data_fim
    )
    emprestimos_periodo_ant = Emprestimo.objects.filter(
        data_emprestimo__gte=data_inicio_ant, data_emprestimo__lte=data_fim_ant
    )

    if turma_filtro:
        emprestimos_periodo = emprestimos_periodo.filter(usuario__serie=turma_filtro)
        emprestimos_periodo_ant = emprestimos_periodo_ant.filter(usuario__serie=turma_filtro)

    # ── 1) Empréstimos no mês ──────────────────────────────
    total_emprestimos = emprestimos_periodo.count()
    total_emprestimos_ant = emprestimos_periodo_ant.count()
    variacao_emprestimos = _variacao_pct(total_emprestimos, total_emprestimos_ant)

    # ── 2) Leitores ativos (>= 1 empréstimo no período) ────
    leitores_ativos = emprestimos_periodo.values('usuario_id').distinct().count()
    leitores_ativos_ant = emprestimos_periodo_ant.values('usuario_id').distinct().count()
    variacao_leitores = leitores_ativos - leitores_ativos_ant

    total_cadastrados_qs = Usuario.objects.filter(tipo_usuario='aluno')
    if turma_filtro:
        total_cadastrados_qs = total_cadastrados_qs.filter(serie=turma_filtro)
    total_cadastrados = total_cadastrados_qs.count()

    # ── 3) Taxa de atraso ───────────────────────────────────
    atrasados_count = _contar_atrasados(emprestimos_periodo, hoje)
    taxa_atraso = round((atrasados_count / total_emprestimos * 100), 1) if total_emprestimos else 0.0

    atrasados_count_ant = _contar_atrasados(emprestimos_periodo_ant, hoje)
    taxa_atraso_ant = round((atrasados_count_ant / total_emprestimos_ant * 100), 1) if total_emprestimos_ant else 0.0
    variacao_atraso = round(taxa_atraso - taxa_atraso_ant, 1)

    # ── 4) Livro mais pedido (reservas -> fallback empréstimos) ──
    reservas_periodo = Reserva.objects.filter(
        data_reserva__date__gte=data_inicio,
        data_reserva__date__lte=data_fim,
        status='pendente',
    )
    if turma_filtro:
        reservas_periodo = reservas_periodo.filter(usuario__serie=turma_filtro)

    usando_reservas = reservas_periodo.exists()

    if usando_reservas:
        row = (
            reservas_periodo.values('livro__titulo')
            .annotate(total=Count('id'))
            .order_by('-total')
            .first()
        )
        livro_mais_pedido = row['livro__titulo'] if row else '—'
        livro_mais_pedido_qtd = row['total'] if row else 0
    else:
        row = (
            emprestimos_periodo.values('exemplar__livro__titulo')
            .annotate(total=Count('id'))
            .order_by('-total')
            .first()
        )
        livro_mais_pedido = row['exemplar__livro__titulo'] if row else '—'
        livro_mais_pedido_qtd = row['total'] if row else 0

    # ── 5) Top 10 livros mais emprestados no período ───────
    top_livros_qs = (
        emprestimos_periodo.values(
            'exemplar__livro__id_livro', 'exemplar__livro__titulo', 'exemplar__livro__capa_url'
        )
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_livros_qs = list(top_livros_qs)
    max_livro_total = top_livros_qs[0]['total'] if top_livros_qs else 1
    top_livros = [{
        'posicao': i + 1,
        'titulo': item['exemplar__livro__titulo'],
        'total': item['total'],
        'percentual': int((item['total'] / max_livro_total) * 100) if max_livro_total else 0,
        'cor': _cor_ranking(i),
        'capa_url': item['exemplar__livro__capa_url'] or None,
    } for i, item in enumerate(top_livros_qs)]

    # ── 6) Top 10 alunos que mais leram no período ─────────
    top_alunos_qs = (
        emprestimos_periodo.values(
            'usuario__id', 'usuario__first_name', 'usuario__last_name', 'usuario__serie'
        )
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_alunos_qs = list(top_alunos_qs)
    max_aluno_total = top_alunos_qs[0]['total'] if top_alunos_qs else 1
    top_alunos = [{
        'posicao': i + 1,
        'nome': f"{item['usuario__first_name']} {item['usuario__last_name']}".strip() or 'Sem nome',
        'turma': item['usuario__serie'] or '—',
        'total': item['total'],
        'percentual': int((item['total'] / max_aluno_total) * 100) if max_aluno_total else 0,
        'cor': _cor_ranking(i),
    } for i, item in enumerate(top_alunos_qs)]

    grafico_dados = _dados_grafico(intervalo, turma_filtro)

    return {
        'periodo_selecionado': periodo_str,
        'periodo_label': f"{MESES_PT[mes]} de {ano}",
        'periodo_label_curto': f"{MESES_PT_ABREV[mes]} {ano}",
        'turma_filtro': turma_filtro,
        'intervalo_selecionado': intervalo,

        'total_emprestimos': total_emprestimos,
        'variacao_emprestimos': variacao_emprestimos,

        'leitores_ativos': leitores_ativos,
        'variacao_leitores': variacao_leitores,
        'total_cadastrados': total_cadastrados,

        'taxa_atraso': taxa_atraso,
        'variacao_atraso': variacao_atraso,
        'atrasados_count': atrasados_count,

        'livro_mais_pedido': livro_mais_pedido,
        'livro_mais_pedido_qtd': livro_mais_pedido_qtd,
        'usando_reservas': usando_reservas,

        'top_livros': top_livros,
        'top_alunos': top_alunos,

        'grafico_labels_json': json.dumps(grafico_dados['labels']),
        'grafico_valores_json': json.dumps(grafico_dados['valores']),
        'grafico_hoje_index': grafico_dados['hoje_index'],
    }


@login_required
def relatorios(request):
    hoje = date.today()
    periodo_str = request.GET.get('mes', hoje.strftime('%Y-%m'))
    turma_filtro = request.GET.get('turma', '').strip()
    intervalo = request.GET.get('intervalo', '30')

    contexto = _montar_contexto_relatorio(periodo_str, turma_filtro, intervalo)
    contexto['turmas'] = Turma.objects.values_list('nome', flat=True)

    return render(request, 'biblioteca/relatorios.html', contexto)


def relatorios_grafico(request):
    """Endpoint AJAX chamado ao trocar o intervalo do gráfico (30d / 90d / 1 ano)."""
    intervalo = request.GET.get('intervalo', '30')
    turma_filtro = request.GET.get('turma', '').strip()
    dados = _dados_grafico(intervalo, turma_filtro)
    return JsonResponse(dados)


@staff_member_required
def exportar_relatorio_pdf(request):
    """Gera o PDF do relatório (paisagem) para a bibliotecária imprimir/colar."""
    hoje = date.today()
    periodo_str = request.GET.get('mes', hoje.strftime('%Y-%m'))
    turma_filtro = request.GET.get('turma', '').strip()

    contexto = _montar_contexto_relatorio(periodo_str, turma_filtro, intervalo='30')
    contexto['gerado_em'] = timezone.now().strftime('%d/%m/%Y às %H:%M')
    contexto['turma_label'] = turma_filtro if turma_filtro else 'Todas as turmas'

    grafico_pdf = _grafico_barras_pdf('30', turma_filtro)
    contexto['grafico_barras'] = grafico_pdf['barras']
    contexto['grafico_valores_max'] = grafico_pdf['maximo']

    html_renderizado = render(request, 'biblioteca/relatorio_pdf.html', contexto).content.decode('utf-8')

    from weasyprint import HTML
    pdf_bytes = HTML(
        string=html_renderizado,
        base_url=request.build_absolute_uri('/'),
    ).write_pdf()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    nome_arquivo = f"Relatorio_Biblioteca_{periodo_str}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response